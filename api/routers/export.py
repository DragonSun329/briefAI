"""
Professional Bulk Export API for briefAI.

Bloomberg-quality export capabilities:
- CSV bulk export with streaming
- JSON Lines (JSONL) streaming export
- Excel export with formatting and multiple sheets
- Parquet for data science workflows
- Paginated API for large datasets
- Async job processing for huge exports
"""

import asyncio
import csv
import io
import json
import uuid
import gzip
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any, Literal, Generator, AsyncIterator
from pathlib import Path
from enum import Enum
import threading
import sqlite3

from fastapi import APIRouter, Query, HTTPException, BackgroundTasks, Depends, Response
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel, Field

import sys
_app_dir = Path(__file__).parent.parent.parent
if str(_app_dir) not in sys.path:
    sys.path.insert(0, str(_app_dir))

from utils.signal_store import SignalStore
from api.auth import verify_api_key, APIKeyInfo, require_feature, require_tier


router = APIRouter(prefix="/api/v1/export", tags=["export"])


# =============================================================================
# Enums and Models
# =============================================================================

class ExportFormat(str, Enum):
    CSV = "csv"
    JSON = "json"
    JSONL = "jsonl"
    EXCEL = "excel"
    PARQUET = "parquet"


class ExportJobStatus(str, Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    EXPIRED = "expired"


class ExportJob(BaseModel):
    """Export job metadata."""
    job_id: str
    export_type: str
    format: ExportFormat
    status: ExportJobStatus
    created_at: datetime
    completed_at: Optional[datetime] = None
    expires_at: Optional[datetime] = None
    file_path: Optional[str] = None
    file_size_bytes: Optional[int] = None
    row_count: Optional[int] = None
    error: Optional[str] = None
    progress: float = 0.0
    download_url: Optional[str] = None


class ExportJobCreate(BaseModel):
    """Request to create an export job."""
    export_type: Literal["signals", "entities", "events", "profiles", "divergences", "articles", "full_dataset"]
    format: ExportFormat = ExportFormat.JSON
    start_date: Optional[str] = Field(None, description="Start date YYYY-MM-DD")
    end_date: Optional[str] = Field(None, description="End date YYYY-MM-DD")
    entity_ids: Optional[List[str]] = Field(None, description="Filter by entity IDs")
    categories: Optional[List[str]] = Field(None, description="Filter by signal categories")
    sectors: Optional[List[str]] = Field(None, description="Filter by sectors")
    min_score: Optional[float] = Field(None, ge=0, le=100, description="Minimum score threshold")
    include_metadata: bool = Field(True, description="Include export metadata in file")
    compress: bool = Field(False, description="GZIP compress the output")
    
    class Config:
        json_schema_extra = {
            "example": {
                "export_type": "signals",
                "format": "csv",
                "start_date": "2025-01-01",
                "end_date": "2025-06-01",
                "categories": ["technical", "financial"],
                "min_score": 50,
                "compress": True
            }
        }


class PaginatedExportRequest(BaseModel):
    """Paginated export request for large datasets."""
    export_type: Literal["signals", "entities", "profiles", "divergences"]
    page: int = Field(1, ge=1)
    page_size: int = Field(1000, ge=100, le=10000)
    format: ExportFormat = ExportFormat.JSON
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    entity_ids: Optional[List[str]] = None
    order_by: str = Field("created_at", description="Field to sort by")
    order_desc: bool = True
    
    class Config:
        json_schema_extra = {
            "example": {
                "export_type": "signals",
                "page": 1,
                "page_size": 5000,
                "format": "jsonl",
                "order_by": "score",
                "order_desc": True
            }
        }


class PaginatedExportResponse(BaseModel):
    """Response for paginated export."""
    page: int
    page_size: int
    total_pages: int
    total_rows: int
    has_next: bool
    has_prev: bool
    data_url: str
    next_page_url: Optional[str] = None


# =============================================================================
# In-memory job store (production: use Redis/PostgreSQL)
# =============================================================================

_export_jobs: Dict[str, ExportJob] = {}
_export_files: Dict[str, bytes] = {}
_jobs_lock = threading.Lock()

_store: Optional[SignalStore] = None


def get_store() -> SignalStore:
    global _store
    if _store is None:
        _store = SignalStore()
    return _store


# =============================================================================
# Synchronous Export (small-medium datasets - streaming)
# =============================================================================

@router.get(
    "/signals",
    summary="Export Signal Scores",
    description="""
Export signal scores in various formats.

**Formats:** CSV, JSON, JSONL, Parquet

**Filters:**
- Date range
- Entity IDs
- Signal categories
- Minimum score

**Example:**
```
GET /api/v1/export/signals?format=csv&start_date=2025-01-01&categories=technical,financial
```
    """,
    responses={
        200: {
            "description": "Export file",
            "content": {
                "text/csv": {},
                "application/json": {},
                "application/x-ndjson": {},
                "application/octet-stream": {},
            }
        }
    }
)
async def export_signals(
    format: ExportFormat = Query(ExportFormat.JSON),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    entity_ids: Optional[str] = Query(None, description="Comma-separated entity IDs"),
    categories: Optional[str] = Query(None, description="Comma-separated categories"),
    min_score: Optional[float] = Query(None, ge=0, le=100),
    limit: int = Query(10000, le=100000),
    offset: int = Query(0, ge=0),
    compress: bool = Query(False, description="GZIP compress output"),
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """Export signal scores with streaming response."""
    store = get_store()
    
    entity_list = entity_ids.split(",") if entity_ids else None
    category_list = categories.split(",") if categories else None
    
    rows = _query_signals(
        store, 
        start_date=start_date,
        end_date=end_date,
        entity_ids=entity_list,
        categories=category_list,
        min_score=min_score,
        limit=limit,
        offset=offset,
    )
    
    return _create_export_response(rows, format, "signals", compress)


@router.get("/signals/stream")
async def stream_signals(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    entity_ids: Optional[str] = Query(None),
    categories: Optional[str] = Query(None),
    key: APIKeyInfo = Depends(require_tier("premium")),
):
    """
    Stream signals as JSON Lines (JSONL) for large datasets.
    
    Premium feature - streams data without buffering entire dataset.
    Ideal for data pipelines and large exports.
    """
    store = get_store()
    
    entity_list = entity_ids.split(",") if entity_ids else None
    category_list = categories.split(",") if categories else None
    
    async def generate() -> AsyncIterator[bytes]:
        conn = store._get_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT * FROM signal_scores WHERE 1=1
        """
        params = []
        
        if start_date:
            query += " AND created_at >= ?"
            params.append(start_date)
        if end_date:
            query += " AND created_at <= ?"
            params.append(end_date + "T23:59:59")
        if entity_list:
            placeholders = ",".join("?" * len(entity_list))
            query += f" AND entity_id IN ({placeholders})"
            params.extend(entity_list)
        if category_list:
            placeholders = ",".join("?" * len(category_list))
            query += f" AND category IN ({placeholders})"
            params.extend(category_list)
        
        query += " ORDER BY created_at DESC"
        
        cursor.execute(query, params)
        
        while True:
            rows = cursor.fetchmany(1000)
            if not rows:
                break
            for row in rows:
                yield json.dumps(dict(row), default=str).encode() + b"\n"
        
        conn.close()
    
    return StreamingResponse(
        generate(),
        media_type="application/x-ndjson",
        headers={
            "Content-Disposition": "attachment; filename=briefai_signals_stream.jsonl",
        },
    )


@router.get("/entities")
async def export_entities(
    format: ExportFormat = Query(ExportFormat.JSON),
    entity_type: Optional[str] = Query(None),
    sector: Optional[str] = Query(None),
    limit: int = Query(10000, le=100000),
    compress: bool = Query(False),
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """Export entity master list."""
    store = get_store()
    rows = _query_entities(store, entity_type=entity_type, sector=sector, limit=limit)
    return _create_export_response(rows, format, "entities", compress)


@router.get("/profiles")
async def export_profiles(
    format: ExportFormat = Query(ExportFormat.JSON),
    entity_type: Optional[str] = Query(None),
    min_score: Optional[float] = Query(None, ge=0, le=100),
    sector: Optional[str] = Query(None),
    limit: int = Query(10000, le=100000),
    compress: bool = Query(False),
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """Export signal profiles."""
    store = get_store()
    rows = _query_profiles(store, entity_type=entity_type, min_score=min_score, sector=sector, limit=limit)
    return _create_export_response(rows, format, "profiles", compress)


@router.get("/divergences")
async def export_divergences(
    format: ExportFormat = Query(ExportFormat.JSON),
    interpretation: Optional[str] = Query(None),
    include_resolved: bool = Query(False),
    min_magnitude: Optional[float] = Query(None, ge=0),
    limit: int = Query(10000, le=100000),
    compress: bool = Query(False),
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """Export signal divergences."""
    store = get_store()
    rows = _query_divergences(
        store, 
        interpretation=interpretation,
        include_resolved=include_resolved,
        min_magnitude=min_magnitude,
        limit=limit,
    )
    return _create_export_response(rows, format, "divergences", compress)


# =============================================================================
# Paginated Export API
# =============================================================================

@router.post("/paginated", response_model=PaginatedExportResponse)
async def paginated_export(
    request: PaginatedExportRequest,
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """
    Paginated export for large datasets.
    
    Returns pagination metadata and a URL to download the current page.
    Use for datasets > 100k rows or when you need controlled iteration.
    
    **Example workflow:**
    1. POST with page=1, page_size=5000
    2. Download data from data_url
    3. Check has_next, use next_page_url for next batch
    4. Repeat until has_next=False
    """
    store = get_store()
    
    # Get total count
    total = _get_total_count(store, request.export_type, request)
    total_pages = (total + request.page_size - 1) // request.page_size
    
    # Calculate offset
    offset = (request.page - 1) * request.page_size
    
    # Generate temporary token for download
    download_token = str(uuid.uuid4())[:8]
    
    base_url = "/api/v1/export"
    data_url = (
        f"{base_url}/{request.export_type}?"
        f"format={request.format.value}&limit={request.page_size}&offset={offset}"
    )
    
    if request.start_date:
        data_url += f"&start_date={request.start_date}"
    if request.end_date:
        data_url += f"&end_date={request.end_date}"
    
    next_page_url = None
    if request.page < total_pages:
        next_page_url = f"{base_url}/paginated"  # Would include query params
    
    return PaginatedExportResponse(
        page=request.page,
        page_size=request.page_size,
        total_pages=total_pages,
        total_rows=total,
        has_next=request.page < total_pages,
        has_prev=request.page > 1,
        data_url=data_url,
        next_page_url=next_page_url,
    )


# =============================================================================
# Async Export Jobs (large datasets)
# =============================================================================

@router.post(
    "/jobs", 
    response_model=ExportJob,
    summary="Create Async Export Job",
    description="""
Create an asynchronous export job for large datasets.

Jobs are processed in the background. Poll the job status endpoint
to check completion, then download the result.

**Use cases:**
- Exporting > 100k rows
- Full dataset snapshots
- Excel exports (always async due to processing)
- Compressed archives

**Job lifecycle:**
1. PENDING → 2. PROCESSING → 3. COMPLETED/FAILED
4. EXPIRED (after 24h)
    """
)
async def create_export_job(
    request: ExportJobCreate,
    background_tasks: BackgroundTasks,
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """Create an async export job for large datasets."""
    job_id = str(uuid.uuid4())
    
    job = ExportJob(
        job_id=job_id,
        export_type=request.export_type,
        format=request.format,
        status=ExportJobStatus.PENDING,
        created_at=datetime.utcnow(),
        expires_at=datetime.utcnow() + timedelta(hours=24),
    )
    
    with _jobs_lock:
        _export_jobs[job_id] = job
    
    background_tasks.add_task(_process_export_job, job_id, request)
    
    return job


@router.get("/jobs/{job_id}", response_model=ExportJob)
async def get_export_job(
    job_id: str,
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """Get status of an export job."""
    with _jobs_lock:
        if job_id not in _export_jobs:
            raise HTTPException(status_code=404, detail="Job not found")
        
        job = _export_jobs[job_id]
        
        # Check expiration
        if job.expires_at and datetime.utcnow() > job.expires_at:
            job.status = ExportJobStatus.EXPIRED
        
        # Add download URL if completed
        if job.status == ExportJobStatus.COMPLETED:
            job.download_url = f"/api/v1/export/jobs/{job_id}/download"
        
        return job


@router.get("/jobs/{job_id}/download")
async def download_export_job(
    job_id: str,
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """Download the result of a completed export job."""
    with _jobs_lock:
        if job_id not in _export_jobs:
            raise HTTPException(status_code=404, detail="Job not found")
        
        job = _export_jobs[job_id]
        
        if job.status == ExportJobStatus.EXPIRED:
            raise HTTPException(status_code=410, detail="Export has expired")
        
        if job.status != ExportJobStatus.COMPLETED:
            raise HTTPException(
                status_code=400, 
                detail=f"Job not ready. Status: {job.status}"
            )
        
        if job_id not in _export_files:
            raise HTTPException(status_code=404, detail="Export file not found")
        
        data = _export_files[job_id]
    
    # Determine content type and filename
    content_type, ext = _get_content_type(job.format)
    filename = f"briefai_{job.export_type}_{job.created_at.strftime('%Y%m%d')}_{job_id[:8]}.{ext}"
    
    return StreamingResponse(
        io.BytesIO(data),
        media_type=content_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(data)),
        },
    )


@router.get("/jobs", response_model=List[ExportJob])
async def list_export_jobs(
    status: Optional[ExportJobStatus] = Query(None),
    limit: int = Query(20, le=100),
    key: APIKeyInfo = Depends(require_feature("export_enabled")),
):
    """List recent export jobs."""
    with _jobs_lock:
        jobs = list(_export_jobs.values())
        
        if status:
            jobs = [j for j in jobs if j.status == status]
        
        jobs = sorted(jobs, key=lambda j: j.created_at, reverse=True)[:limit]
    
    return jobs


@router.delete("/jobs/{job_id}")
async def delete_export_job(
    job_id: str,
    key: APIKeyInfo = Depends(require_tier("premium")),
):
    """Delete an export job and its data."""
    with _jobs_lock:
        if job_id not in _export_jobs:
            raise HTTPException(status_code=404, detail="Job not found")
        
        del _export_jobs[job_id]
        if job_id in _export_files:
            del _export_files[job_id]
    
    return {"status": "deleted", "job_id": job_id}


# =============================================================================
# Excel Export (Premium)
# =============================================================================

@router.post("/excel", response_model=ExportJob)
async def create_excel_export(
    export_types: List[str] = Query(
        ["signals", "entities", "profiles"],
        description="Tables to include as sheets"
    ),
    include_charts: bool = Query(False, description="Include summary charts"),
    background_tasks: BackgroundTasks = None,
    key: APIKeyInfo = Depends(require_tier("premium")),
):
    """
    Create a formatted Excel export with multiple sheets.
    
    Premium feature. Includes:
    - Multiple data sheets (one per export type)
    - Formatted headers and column widths
    - Optional summary charts
    - Metadata sheet with export info
    """
    job_id = str(uuid.uuid4())
    
    job = ExportJob(
        job_id=job_id,
        export_type="excel_multi",
        format=ExportFormat.EXCEL,
        status=ExportJobStatus.PENDING,
        created_at=datetime.utcnow(),
        expires_at=datetime.utcnow() + timedelta(hours=24),
    )
    
    with _jobs_lock:
        _export_jobs[job_id] = job
    
    background_tasks.add_task(_process_excel_export, job_id, export_types, include_charts)
    
    return job


# =============================================================================
# Helper Functions
# =============================================================================

def _get_content_type(format: ExportFormat) -> tuple[str, str]:
    """Get content type and file extension for format."""
    content_types = {
        ExportFormat.CSV: ("text/csv", "csv"),
        ExportFormat.JSON: ("application/json", "json"),
        ExportFormat.JSONL: ("application/x-ndjson", "jsonl"),
        ExportFormat.EXCEL: ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
        ExportFormat.PARQUET: ("application/octet-stream", "parquet"),
    }
    return content_types.get(format, ("application/octet-stream", "bin"))


def _get_total_count(store: SignalStore, export_type: str, request: PaginatedExportRequest) -> int:
    """Get total count for pagination."""
    conn = store._get_connection()
    cursor = conn.cursor()
    
    table_map = {
        "signals": "signal_scores",
        "entities": "entities",
        "profiles": "signal_profiles",
        "divergences": "signal_divergences",
    }
    
    table = table_map.get(export_type, "signal_scores")
    query = f"SELECT COUNT(*) FROM {table} WHERE 1=1"
    params = []
    
    if request.start_date:
        query += " AND created_at >= ?"
        params.append(request.start_date)
    if request.end_date:
        query += " AND created_at <= ?"
        params.append(request.end_date + "T23:59:59")
    
    cursor.execute(query, params)
    count = cursor.fetchone()[0]
    conn.close()
    
    return count


def _query_signals(
    store: SignalStore,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    entity_ids: Optional[List[str]] = None,
    categories: Optional[List[str]] = None,
    min_score: Optional[float] = None,
    limit: int = 10000,
    offset: int = 0,
) -> List[Dict[str, Any]]:
    """Query signal scores with filters."""
    conn = store._get_connection()
    cursor = conn.cursor()
    
    query = """
        SELECT 
            s.id, s.entity_id, s.source_id, s.category,
            s.score, s.percentile, s.score_delta_7d, s.score_delta_30d,
            s.period_start, s.period_end, s.created_at
        FROM signal_scores s
        WHERE 1=1
    """
    params = []
    
    if start_date:
        query += " AND s.created_at >= ?"
        params.append(start_date)
    if end_date:
        query += " AND s.created_at <= ?"
        params.append(end_date + "T23:59:59")
    if entity_ids:
        placeholders = ",".join("?" * len(entity_ids))
        query += f" AND s.entity_id IN ({placeholders})"
        params.extend(entity_ids)
    if categories:
        placeholders = ",".join("?" * len(categories))
        query += f" AND s.category IN ({placeholders})"
        params.extend(categories)
    if min_score is not None:
        query += " AND s.score >= ?"
        params.append(min_score)
    
    query += f" ORDER BY s.created_at DESC LIMIT {limit} OFFSET {offset}"
    
    cursor.execute(query, params)
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return rows


def _query_entities(
    store: SignalStore,
    entity_type: Optional[str] = None,
    sector: Optional[str] = None,
    limit: int = 10000,
) -> List[Dict[str, Any]]:
    """Query entities with filters."""
    conn = store._get_connection()
    cursor = conn.cursor()
    
    query = "SELECT * FROM entities WHERE 1=1"
    params = []
    
    if entity_type:
        query += " AND entity_type = ?"
        params.append(entity_type)
    if sector:
        query += " AND sector = ?"
        params.append(sector)
    
    query += f" ORDER BY name LIMIT {limit}"
    
    cursor.execute(query, params)
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return rows


def _query_profiles(
    store: SignalStore,
    entity_type: Optional[str] = None,
    min_score: Optional[float] = None,
    sector: Optional[str] = None,
    limit: int = 10000,
) -> List[Dict[str, Any]]:
    """Query signal profiles."""
    conn = store._get_connection()
    cursor = conn.cursor()
    
    query = """
        SELECT * FROM signal_profiles p1
        WHERE as_of = (
            SELECT MAX(as_of) FROM signal_profiles p2
            WHERE p2.entity_id = p1.entity_id
        )
    """
    params = []
    
    if entity_type:
        query += " AND entity_type = ?"
        params.append(entity_type)
    if min_score is not None:
        query += " AND composite_score >= ?"
        params.append(min_score)
    if sector:
        query += " AND sector = ?"
        params.append(sector)
    
    query += f" ORDER BY composite_score DESC LIMIT {limit}"
    
    cursor.execute(query, params)
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return rows


def _query_divergences(
    store: SignalStore,
    interpretation: Optional[str] = None,
    include_resolved: bool = False,
    min_magnitude: Optional[float] = None,
    limit: int = 10000,
) -> List[Dict[str, Any]]:
    """Query signal divergences."""
    conn = store._get_connection()
    cursor = conn.cursor()
    
    query = "SELECT * FROM signal_divergences WHERE 1=1"
    params = []
    
    if not include_resolved:
        query += " AND resolved_at IS NULL"
    if interpretation:
        query += " AND interpretation = ?"
        params.append(interpretation)
    if min_magnitude is not None:
        query += " AND divergence_magnitude >= ?"
        params.append(min_magnitude)
    
    query += f" ORDER BY detected_at DESC LIMIT {limit}"
    
    cursor.execute(query, params)
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return rows


def _create_export_response(
    rows: List[Dict[str, Any]],
    format: ExportFormat,
    export_name: str,
    compress: bool = False,
) -> StreamingResponse:
    """Create a streaming export response."""
    
    if format == ExportFormat.JSON:
        content = json.dumps({
            "meta": {
                "export_type": export_name,
                "row_count": len(rows),
                "exported_at": datetime.utcnow().isoformat(),
            },
            "data": rows,
        }, default=str, indent=2)
        media_type = "application/json"
        ext = "json"
    
    elif format == ExportFormat.JSONL:
        lines = [json.dumps(row, default=str) for row in rows]
        content = "\n".join(lines)
        media_type = "application/x-ndjson"
        ext = "jsonl"
    
    elif format == ExportFormat.CSV:
        if not rows:
            content = ""
        else:
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
            content = output.getvalue()
        media_type = "text/csv"
        ext = "csv"
    
    elif format == ExportFormat.PARQUET:
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
            
            if rows:
                table = pa.Table.from_pylist(rows)
                buffer = io.BytesIO()
                pq.write_table(table, buffer)
                content = buffer.getvalue()
            else:
                content = b""
            
            if compress:
                content = gzip.compress(content)
            
            return StreamingResponse(
                io.BytesIO(content),
                media_type="application/octet-stream",
                headers={
                    "Content-Disposition": f"attachment; filename=briefai_{export_name}.parquet" + (".gz" if compress else ""),
                },
            )
        except ImportError:
            raise HTTPException(
                status_code=501,
                detail="Parquet export requires pyarrow. Install with: pip install pyarrow",
            )
    
    elif format == ExportFormat.EXCEL:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = export_name.title()
            
            if rows:
                # Header row
                headers = list(rows[0].keys())
                header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Data rows
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row.get(header)
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 15
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            content = buffer.getvalue()
            
            return StreamingResponse(
                io.BytesIO(content),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=briefai_{export_name}.xlsx",
                },
            )
        except ImportError:
            raise HTTPException(
                status_code=501,
                detail="Excel export requires openpyxl. Install with: pip install openpyxl",
            )
    
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {format}")
    
    # Apply compression if requested
    if compress and isinstance(content, str):
        content = gzip.compress(content.encode())
        ext += ".gz"
        media_type = "application/gzip"
    elif isinstance(content, str):
        content = content.encode()
    
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename=briefai_{export_name}.{ext}",
        },
    )


async def _process_export_job(job_id: str, request: ExportJobCreate):
    """Process an export job in the background."""
    store = get_store()
    
    with _jobs_lock:
        if job_id not in _export_jobs:
            return
        _export_jobs[job_id].status = ExportJobStatus.PROCESSING
        _export_jobs[job_id].progress = 0.1
    
    try:
        # Query data based on export type
        if request.export_type == "signals":
            rows = _query_signals(
                store,
                start_date=request.start_date,
                end_date=request.end_date,
                entity_ids=request.entity_ids,
                categories=request.categories,
                min_score=request.min_score,
                limit=1000000,
            )
        elif request.export_type == "entities":
            rows = _query_entities(store, limit=1000000)
        elif request.export_type == "profiles":
            rows = _query_profiles(store, min_score=request.min_score, limit=1000000)
        elif request.export_type == "divergences":
            rows = _query_divergences(store, include_resolved=True, limit=1000000)
        elif request.export_type == "full_dataset":
            # Export all tables
            rows = {
                "signals": _query_signals(store, limit=1000000),
                "entities": _query_entities(store, limit=1000000),
                "profiles": _query_profiles(store, limit=1000000),
                "divergences": _query_divergences(store, include_resolved=True, limit=1000000),
            }
        else:
            raise ValueError(f"Unknown export type: {request.export_type}")
        
        with _jobs_lock:
            _export_jobs[job_id].progress = 0.5
        
        # Convert to requested format
        if request.export_type == "full_dataset":
            # Multi-table export to JSON
            data = json.dumps({
                "meta": {
                    "export_type": "full_dataset",
                    "exported_at": datetime.utcnow().isoformat(),
                    "tables": list(rows.keys()),
                },
                **rows,
            }, default=str, indent=2).encode()
        elif request.format == ExportFormat.JSON:
            data = json.dumps({
                "meta": {
                    "export_type": request.export_type,
                    "row_count": len(rows),
                    "exported_at": datetime.utcnow().isoformat(),
                },
                "data": rows,
            }, default=str, indent=2).encode()
        elif request.format == ExportFormat.JSONL:
            lines = [json.dumps(row, default=str) for row in rows]
            data = "\n".join(lines).encode()
        elif request.format == ExportFormat.CSV:
            output = io.StringIO()
            if rows:
                writer = csv.DictWriter(output, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
            data = output.getvalue().encode()
        elif request.format == ExportFormat.PARQUET:
            try:
                import pyarrow as pa
                import pyarrow.parquet as pq
                
                if rows:
                    table = pa.Table.from_pylist(rows)
                    buffer = io.BytesIO()
                    pq.write_table(table, buffer)
                    data = buffer.getvalue()
                else:
                    data = b""
            except ImportError:
                raise ValueError("Parquet export requires pyarrow")
        elif request.format == ExportFormat.EXCEL:
            data = await _create_excel_bytes(rows, request.export_type)
        else:
            raise ValueError(f"Unknown format: {request.format}")
        
        # Compress if requested
        if request.compress:
            data = gzip.compress(data)
        
        row_count = len(rows) if isinstance(rows, list) else sum(len(v) for v in rows.values())
        
        with _jobs_lock:
            _export_jobs[job_id].status = ExportJobStatus.COMPLETED
            _export_jobs[job_id].completed_at = datetime.utcnow()
            _export_jobs[job_id].row_count = row_count
            _export_jobs[job_id].file_size_bytes = len(data)
            _export_jobs[job_id].progress = 1.0
            _export_files[job_id] = data
    
    except Exception as e:
        with _jobs_lock:
            _export_jobs[job_id].status = ExportJobStatus.FAILED
            _export_jobs[job_id].error = str(e)


async def _create_excel_bytes(rows: List[Dict], sheet_name: str) -> bytes:
    """Create Excel file bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name.title()[:31]  # Excel sheet name limit
        
        if rows:
            headers = list(rows[0].keys())
            header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header)
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    except ImportError:
        raise ValueError("Excel export requires openpyxl")


async def _process_excel_export(job_id: str, export_types: List[str], include_charts: bool):
    """Process multi-sheet Excel export."""
    store = get_store()
    
    with _jobs_lock:
        if job_id not in _export_jobs:
            return
        _export_jobs[job_id].status = ExportJobStatus.PROCESSING
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Add metadata sheet
        meta_ws = wb.create_sheet("Export Info")
        meta_ws["A1"] = "briefAI Data Export"
        meta_ws["A1"].font = Font(bold=True, size=14)
        meta_ws["A3"] = "Export Date:"
        meta_ws["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        meta_ws["A4"] = "Included Tables:"
        meta_ws["B4"] = ", ".join(export_types)
        
        header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        total_rows = 0
        
        for export_type in export_types:
            if export_type == "signals":
                rows = _query_signals(store, limit=100000)
            elif export_type == "entities":
                rows = _query_entities(store, limit=100000)
            elif export_type == "profiles":
                rows = _query_profiles(store, limit=100000)
            elif export_type == "divergences":
                rows = _query_divergences(store, include_resolved=True, limit=100000)
            else:
                continue
            
            ws = wb.create_sheet(export_type.title())
            
            if rows:
                headers = list(rows[0].keys())
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row.get(header)
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 15
                
                total_rows += len(rows)
        
        # Update metadata
        meta_ws["A5"] = "Total Rows:"
        meta_ws["B5"] = total_rows
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        data = buffer.getvalue()
        
        with _jobs_lock:
            _export_jobs[job_id].status = ExportJobStatus.COMPLETED
            _export_jobs[job_id].completed_at = datetime.utcnow()
            _export_jobs[job_id].row_count = total_rows
            _export_jobs[job_id].file_size_bytes = len(data)
            _export_jobs[job_id].progress = 1.0
            _export_files[job_id] = data
    
    except Exception as e:
        with _jobs_lock:
            _export_jobs[job_id].status = ExportJobStatus.FAILED
            _export_jobs[job_id].error = str(e)
